import openpyxl
from pprint import pprint

filename = 'Daily Report December.xlsx'
month = filename.split('.xlsx')[0].split(' ')[1]
wb_data = openpyxl.load_workbook(filename, data_only=True)

sheetnames = wb_data.get_sheet_names()

items = []
for sheetname in sheetnames:
	sheet = wb_data.get_sheet_by_name(sheetname)

	# Reading values
	location = sheet['B1']
	date = sheet['B2']
	print(date.value)

	starting_item_col = 'B'
	current_item_row = 10
	while sheet['%s%d' % (starting_item_col, current_item_row)].value:
		item = {}
		item['date'] = date.value.strftime("%m/%d/%y")
		item['reference_num'] = sheet['%s%d' % (starting_item_col, current_item_row)].value.upper()
		item['retail_s_price'] = sheet['%s%d' % ('D', current_item_row)].value
		item['net_price'] = sheet['%s%d' % ('E', current_item_row)].value
		item['diff']  = sheet['%s%d' % ('F', current_item_row)].value

		# Determine payment mode

		pprint(item)
		items.append(item)
		current_item_row += 1
pprint(items)

"""##########################################################################################"""

outfilename = 'BP Monthly Report Template.xlsx'
out_wb_data = openpyxl.load_workbook(outfilename)
sheet = out_wb_data.get_sheet_by_name('Sales')
current_item_row = 12

for item in items:
	print("Processing %s" % item['reference_num'])
	sheet['A%d' % current_item_row] = item['reference_num']
	sheet['C%d' % current_item_row] = item['date']
	sheet['D%d' % current_item_row] = 1
	sheet['E%d' % current_item_row] = item['net_price']
	# Determine if discount was provided
	if item['net_price'] != item['retail_s_price']:
		sheet['F%d' % current_item_row] = item['retail_s_price']
	# Determine Sale Type
	if item['reference_num'].startswith('N'):
		sheet['G%d' % current_item_row] = "Watch"
		if '#' in item['reference_num']:
			cols = item['reference_num'].split('#')
			sheet['A%d' % current_item_row] = cols[0].strip()
			sheet['B%d' % current_item_row] = cols[1].strip()
		elif ' ' in item['reference_num']:
			cols = item['reference_num'].split(' ')
			sheet['A%d' % current_item_row] = cols[0].strip()
			sheet['B%d' % current_item_row] = cols[1].strip()
	elif "REPAIR" in item['reference_num']:
		sheet['G%d' % current_item_row] = "Service"
	else:
		sheet['G%d' % current_item_row] = "Accessory"
	current_item_row += 1
out_wb_data.save('BP %s Report.xlsx' % month)
